import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import BarChart, Reference
import logging
from typing import Dict, List
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
import requests
import random
import time

class ExcelReportGenerator:
    def __init__(self):
        self.logger = self._setup_logger()
        # self.logger.info(f"Starting orgnization of ")
        self.stats = {"sheets": 0, "charts": 0, "backups": 0}
    def _setup_logger(self) -> logging.Logger:
        "The Logging Configuration"
        logger = logging.getLogger(__name__)
        logger.setLevel(logging.INFO) # logging level info and higher order
        handler = logging.FileHandler("excel_automation.log")
        handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")) # formatting log output timestamp - level - message
        console_handler = logging.StreamHandler()
        console_handler.setFormatter(logging.Formatter("%(levelname)s - %(message)s"))
        logger.addHandler(handler)
        logger.addHandler(console_handler)
        return logger
    
    def fetch_data(self) -> pd.DataFrame:
        # return pd.DataFrame({
        #     "Product": ["Widget A", "Widget B", "Widget C"],
        #     "Q1 Sales": [45000, 56000, 89000],
        #     "Q2 Sales": [51000, 63000, 92000]
        # })
        """Fetch data from fakestore API"""
        # logic with manual retries
        max_retries = 5
        backoff_time = 2 # Start with 2 seconds
        for attempt in range(max_retries):
            try:
                url = "https://fakestoreapi.com/products"
                response = requests.get(url, timeout=5)
                response.raise_for_status()
                return pd.DataFrame(response.json())
            except requests.exceptions.RequestException as e:
                self.logger.warning(f"Attempt {attempt + 1} of API call failed: {str(e)}")
                if attempt < max_retries -1 :
                    time.sleep(backoff_time)
                    backoff_time *= 2 # Exponential backoff
                else:
                    self.logger.error("Max retries reached. API call failed.")
                    raise
    
    def _flatten_dict_columns(self) -> pd.DataFrame:
        df = self.fetch_data()
        for col in df.columns:
            # Check if the first row is a dictionary
            if isinstance(df[col].iloc[0], dict):  
                # Expand the dictionary into separate columns
                dict_df = pd.json_normalize(df[col])
                dict_df.columns = [f"{col}_{subcol}" for subcol in dict_df.columns]  # Rename columns
                df = df.drop(columns=[col]).join(dict_df)  # Drop the original column and join expanded columns
                # Add random sales data column
                df["Q1_sales"] = [random.randint(1000, 10000) for _ in range(len(df))]
                df["Q2_sales"] = [random.randint(1000, 10000) for _ in range(len(df))]
                # Calculate Q2 vs Q1 growth and add as a new column
                df["Q2 vs Q1 Growth"] = ((df["Q2_sales"] - df["Q1_sales"]) / df["Q1_sales"] * 100).round(1) 
        return df

    def create_report(self, output_path: Path) -> None:
        "Generate Excel report"
        try:
            data = self._flatten_dict_columns()
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                data.to_excel(writer,sheet_name="Sales Data", index=False)
                self.stats["sheets"] += 1
                #Summary sheet
                workbook = writer.book
                self._add_summary_sheet(workbook, data)
            self.logger.info(f"Report generated: {output_path}")
        except Exception as e:
            self.logger.error(f"Report failed: {str(e)}")
            raise
    # summary sheet of excel workbook
    def _add_summary_sheet(self, workbook : Workbook, data: pd.DataFrame) -> None:
        "Add formatted summary with chart"
        ws = workbook.create_sheet("Executive Summary")
        # Header
        ws["A1"] = "Quarterly Sales Report"
        ws["A1"].font = Font(bold= True, size=14)
        # for row in data.itertuples():
        #     ws.append(row[1:])
        for i, row in enumerate(data.itertuples(index=False), start=1):
            ws.append(row)
            # Apply fill color
            fill_color = PatternFill(start_color="D9EAD3" if i % 2 == 0 else "FCE4D6", end_color="D9EAD3" if i % 2 == 0 else "FCE4D6", fill_type="solid")
            for cell in ws[i]:
                cell.fill = fill_color
        chart = BarChart()
        chart.title = "Product performance report"
        chart.style = 13  # Choose a built-in chart style
        # chart.style = 4
        # values = Reference(ws, min_col=2, max_col=ws.max_column, min_row=1, max_row=ws.max_row)
        # # categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

        # chart.add_data(values, titles_from_data=True)
        # # chart.set_categories(categories)
        # ws.add_chart(chart, "E2")
        chart.x_axis.title = "Products"
        chart.y_axis.title = "Sales Data"
        
        # Reference data for chart
        product_range = Reference(ws, min_col=2, min_row=2, max_row=7)  # Product names
        sales_data_range = Reference(ws, min_col=9, min_row=2, max_row=7)  # sales_data
        # quarter_data_range = Reference(ws, min_col=10, min_row=2, max_row=len(data) + 2)  # quarter_no (Column J)
        chart.add_data(sales_data_range, titles_from_data=False)
        chart.set_categories(product_range)
        
        # Place the chart on the worksheet
        ws.add_chart(chart, "N5")
        self.stats["charts"] += 1
if __name__ == "__main__":
    try:
        generator = ExcelReportGenerator()
        # Generate investor-ready report
        report_path = Path("Sales_Report.xlsx")
        generator.create_report(report_path)   
        # Log performance metrics
        generator.logger.info(f"Charts generated: {generator.stats['charts']}")
    except PermissionError:
        generator.logger.error(f"Error: The file {report_path} is already open. Please close it and try again.")
    except Exception as e:
        generator.logger.error(f"An expeted error occured {e}")


